import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import os
import time

# Caminho para o arquivo Excel
arquivo_excel = "C:\\Users\\BG-PROVISORIO\\Desktop\\Vendas\\Mochilas\\Reserva.xlsx"
aba_nova = "Caracteristicas_Extraidas"

# Carregar o Excel em um DataFrame
df = pd.read_excel(arquivo_excel)

# Especificar o nome da coluna que contém os links
coluna_links = 'Summary_URL'

# Configuração do Selenium WebDriver
chrome_options = Options()
chrome_options.add_argument("--headless")  # Executa o navegador em segundo plano

# Inicia o navegador
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

# Função para extrair a cor e composição do produto
def extrair_informacoes(url):
    try:
        # Abrir a página
        driver.get(url)
        time.sleep(2)  # Aguarda o carregamento da página

        # Extrair a cor
        try:
            cor_element = driver.find_element(By.XPATH, "//div[contains(@class, 'sku-selector-custom__selected')]/span[@class='sku-selector-custom__selected-value']")
            cor = cor_element.text.strip()
        except:
            cor = "Cor não encontrada"

        # Extrair a composição
        try:
            composicao_element = driver.find_element(By.XPATH, "//span[@data-specification-name='Composição']/following-sibling::span")
            composicao = composicao_element.text.strip()
        except:
            composicao = "Composição não encontrada"

        return [cor, composicao]

    except Exception as e:
        return [f"Erro ao acessar {url}: {e}", ""]

# Lista para armazenar os resultados
resultados = []

# Iterar sobre os links no DataFrame e extrair as informações
for link in df[coluna_links]:
    info = extrair_informacoes(link)
    resultados.append([link, 'Cor', info[0]])
    resultados.append([link, 'Composição', info[1]])

# Fechar o navegador
driver.quit()

# Carregar o arquivo Excel existente
if os.path.exists(arquivo_excel):
    wb = load_workbook(arquivo_excel)
    if aba_nova in wb.sheetnames:
        ws = wb[aba_nova]
    else:
        ws = wb.create_sheet(aba_nova)

    # Adicionar cabeçalhos se for a primeira vez inserindo dados
    if ws.max_row == 1:
        ws.append(["Link", "Tipo", "Informação"])

    # Adicionar dados
    for resultado in resultados:
        # Verifica se resultado é uma lista não vazia
        if isinstance(resultado, list) and any(resultado):
            ws.append(resultado)

    # Salvar o arquivo Excel com a nova aba
    wb.save(arquivo_excel)
else:
    print(f"O arquivo {arquivo_excel} não existe.")
