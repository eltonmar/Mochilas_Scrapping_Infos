import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import os

# Caminho para o arquivo Excel
arquivo_excel = "C:\\Users\\BG-PROVISORIO\\Desktop\\Vendas\\Mochilas\\Sestini.xlsx"
aba_nova = "Caracteristicas_Extraidas"

# Carregar o Excel em um DataFrame
df = pd.read_excel(arquivo_excel)

# Especificar o nome da coluna que contém os links
coluna_links = 'Title_URL'

# Configuração do Selenium WebDriver
chrome_options = Options()
chrome_options.add_argument("--headless")  # Executa o navegador em segundo plano

# Inicia o navegador
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

# Função para extrair o conteúdo dos spans com rv-html
def extrair_spans(url):
    try:
        # Abrir a página
        driver.get(url)

        # Esperar até que o elemento da tabela de características esteja presente
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[@class='x-product__specifications-characteristics']"))
        )

        # Capturar todas as linhas da tabela
        linhas = driver.find_elements(By.XPATH, "//div[@class='x-product__specifications-characteristics-content']//tr")

        resultado = []
        for linha in linhas:
            # Extrair o rótulo (categoria) e o valor de cada linha
            label = linha.find_element(By.XPATH, ".//td[@class='label']").text
            valor = linha.find_element(By.XPATH, ".//td[2]/span").text
            resultado.append(f"{label}: {valor}")

        return resultado if resultado else ["Nenhuma característica encontrada"]

    except Exception as e:
        return [f"Erro ao acessar {url}: {e}"]

# Lista para armazenar os resultados
resultados = []

# Iterar sobre os links no Excel e extrair o conteúdo dos spans
for link in df[coluna_links]:
    if pd.notna(link):  # Verificar se o valor não é nulo
        spans_info = extrair_spans(link)
        for info in spans_info:
            resultados.append([link, info])

# Fechar o navegador
driver.quit()

# Carregar o arquivo Excel existente
if os.path.exists(arquivo_excel):
    wb = load_workbook(arquivo_excel)
    if aba_nova in wb.sheetnames:
        ws = wb[aba_nova]
    else:
        ws = wb.create_sheet(aba_nova)

    # Adicionar cabeçalhos, se a aba estiver vazia
    if ws.max_row == 1:
        ws.append(["Link", "Conteúdo do span rv-html"])

    # Adicionar dados
    for resultado in resultados:
        ws.append(resultado)

    # Salvar o arquivo Excel com a nova aba
    wb.save(arquivo_excel)
else:
    print(f"O arquivo {arquivo_excel} não existe.")
