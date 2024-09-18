import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import os

# Caminho para o arquivo Excel
arquivo_excel = "C:\\Users\\BG-PROVISORIO\\Desktop\\Vendas\\Mochilas\\LePostiche.xlsx"
aba_nova = "Caracteristicas_Extraidas"

# Carregar o Excel em um DataFrame
df = pd.read_excel(arquivo_excel)

# Especificar o nome da coluna que contém os links
coluna_links = 'Title_URL'

# Configuração do Selenium WebDriver
chrome_options = Options()
chrome_options.add_argument("--headless")  # Executa o navegador em segundo plano

# Inicia o navegador
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

# Função para extrair o conteúdo das informações
# Função para extrair o conteúdo das informações
def extrair_informacoes(url):
    try:
        # Abrir a página
        driver.get(url)

        # Encontrar todas as divs com a classe 'desc'
        descricoes = driver.find_elements(By.XPATH, "//div[@class='desc']")

        # Verificar se há pelo menos dois elementos com a classe 'desc'
        if len(descricoes) > 1:
            descricao = descricoes[1].get_attribute('innerHTML')  # Pegar o segundo 'desc'
        else:
            return ["Segunda descrição não encontrada"]

        # Processar a descrição para extrair informações relevantes
        resultado = []
        secoes = descricao.split('<strong>')  # Divide as seções

        for secao in secoes:
            if '</strong>' in secao:
                titulo, conteudo = secao.split('</strong>', 1)
                resultado.append(f"{titulo.strip()}: {conteudo.strip()}")
            else:
                resultado.append(secao.strip())

        return resultado if resultado else ["Nenhuma informação encontrada"]
    except Exception as e:
        return [f"Erro ao acessar {url}: {e}"]

# Lista para armazenar os resultados
resultados = []

# Iterar sobre os links no Excel e extrair as informações
for link in df[coluna_links]:
    if pd.notna(link):  # Verificar se o valor não é nulo
        informacoes = extrair_informacoes(link)
        for info in informacoes:
            resultados.append([link, info])

# Fechar o navegador
driver.quit()

# Carregar o arquivo Excel existente
if os.path.exists(arquivo_excel):
    wb = load_workbook(arquivo_excel)

    # Verificar se a aba existe e carregar, ou criar nova
    if aba_nova in wb.sheetnames:
        ws = wb[aba_nova]
    else:
        ws = wb.create_sheet(aba_nova)

    # Adicionar cabeçalhos somente se a planilha estiver vazia
    if ws.max_row == 1:
        ws.append(["Link", "Informação Extraída"])

    # Adicionar dados
    for resultado in resultados:
        ws.append(resultado)

    # Salvar o arquivo Excel com a nova aba
    wb.save(arquivo_excel)
else:
    print(f"O arquivo {arquivo_excel} não existe.")
